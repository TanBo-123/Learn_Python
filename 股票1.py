#!/usr/bin/env python3
# -*- coding:utf-8 -*-

import openpyxl
import matplotlib.pyplot as plt
import numpy as np

#用于matplotlib显示中文
plt.rcParams['font.sans-serif'] = ['SimHei']

fpath =  './股票.xlsx'
stock_SIndex_col = 2 #B
stock_SIndex_row = 2
stock_EIndex_row = 21 #股票数
stock_wave_SIndex_col = 5 #E 波动起始位子
stock_wave_EIndex_col = 34 #波动结束位子
RECODE_CYCLE = 30

#with openpyxl.load_workbook(fpath) as wb:
wb = openpyxl.load_workbook(fpath)
sheets = wb.sheetnames
#print(sheets, type(sheets))

ws = wb[wb.sheetnames[1]]
print(type(ws.cell(stock_SIndex_col,stock_SIndex_row).value))
#一行打印一次
for column,num in zip(ws.iter_rows(min_row=stock_SIndex_row+7,max_row=stock_SIndex_row+7,
                           min_col=stock_wave_SIndex_col, max_col = stock_wave_EIndex_col, values_only=True),
                       range(stock_EIndex_row)):
    plt.figure(num=3,figsize=(10,5))
    LL = []
    for i in range(RECODE_CYCLE):
        if not column[i] == None:
            #print(type(column[i]))
            if type(column[i]) == int or type(column[i]) == float:
                #print(column[i])
                LL.append(column[i])
            elif(type(column[i]) == str):
                LL.append(column[i].split('/')[0])
                #print(LL)
            else:print(type(column[i]),column[i])
        else:
            LL.append(0)
        pass
    i = np.arange(0,RECODE_CYCLE)
    LL = list(map(float,LL))
    print(type(LL),LL)
    print(type(1),i)
    plt.xlabel("时间(天)")
    plt.ylabel('波幅')
    plt.plot(i, LL, color='red', linewidth=1.0, linestyle='-',label = ws.cell(stock_SIndex_col,num+stock_SIndex_row).value)

    plt.legend()
    plt.show()
    pass


#wb.save(fpath)  #read_only state not need